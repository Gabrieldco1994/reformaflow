#!/usr/bin/env python3
"""
Reconciliação financeira do projeto PESSOAL (ReformaFlow) — export de produção vs. master.

Valida se os números do app em produção batem com a planilha-master
(`Consolidacao_Financeira.xlsx`), seguindo a lógica do `LOGICA_CONSOLIDACAO_FINANCEIRA.md` (§10).
Desconsidera lançamentos MANUAIS (duplicam o extrato) e trata pagamento de fatura como neutro.

USO:
    pip install openpyxl
    python reconcile.py --prod reformaflow-pessoal-prod-AAAAMMDD.xlsx

    # opcional: sobrescrever saldo inicial / referência do banco / alvos via master
    python reconcile.py --prod export.xlsx --opening 14285.97 --opening-date 2025-12-31 --bank 17229.88
    python reconcile.py --prod export.xlsx --master Consolidacao_Financeira.xlsx   # recalcula alvos do master

SAÍDA: tabela "calculado (prod) vs alvo (master)" com PASS/FAIL por métrica, + checagens de sanidade.
"""
import argparse
import re
import sys
from collections import Counter

try:
    import openpyxl
except ImportError:
    sys.exit("Falta a lib openpyxl. Rode:  pip install openpyxl")

REALIZED = {"PAGO", "EM_CAIXA"}              # status realizados (já saiu/entrou)
NEUTRAL_CAT = "PAGAMENTO_FATURA_CARTAO"      # categoria neutra (não é gasto real)
CONTA = "Conta Corrente"
MANUAL = "Manual"
RESGATE_RE = re.compile(r"\bRESG(ATE)?\b", re.I)
EST_RESG_RE = re.compile(r"EST\s+RESG", re.I)

# Alvos do master (corte 01/06/2026), em reais. Sobrescritos por --master se informado.
DEFAULT_TARGETS = {
    "recebimentos_reais": 497689.46,
    "saidas_reais": -530954.94,
    "neutras": -76475.64,
    "liquido_realizado": 18381.48,
    "saldo_projetado": -33265.48,
    "caixa_conta": 17229.74,
}
DEFAULT_OPENING = 14285.97
DEFAULT_BANK = 17229.88
DEFAULT_PROJECT = "Controle Financeiro Gab"
TOL = 1.00  # tolerância R$ 1,00

# Query que reproduz as transações do PESSOAL direto do SQLite (modo --db).
# Espelha export_prod.sql. Datas no Prisma/SQLite são INTEGER (epoch ms).
DB_QUERY = """
SELECT
  p.name AS Projeto,
  date(COALESCE(e.data_pagamento, e.data_inicio_parcela, e.created_at) / 1000, 'unixepoch') AS Data,
  CASE
    WHEN e.bank_last4 IS NOT NULL THEN 'Conta Corrente'
    WHEN e.card_last4 IS NOT NULL THEN 'Cartao de Credito'
    ELSE 'Manual'
  END AS Origem,
  COALESCE(e.fornecedor, e.titulo, '') AS Fonte,
  COALESCE(e.titulo, e.fornecedor, '') AS Descricao,
  ROUND(-e.valor_total / 100.0, 2) AS Valor,
  'Saida' AS Tipo,
  e.tipo_despesa AS Categoria,
  e.status AS Status
FROM expenses e JOIN projects p ON p.id = e.project_id
WHERE e.deleted_at IS NULL AND e.linked_expense_id IS NULL AND p.name = ?
UNION ALL
SELECT
  p.name,
  date(r.data / 1000, 'unixepoch'),
  CASE WHEN r.bank_last4 IS NOT NULL THEN 'Conta Corrente' ELSE 'Manual' END,
  COALESCE(r.descricao, ''),
  COALESCE(r.descricao, ''),
  ROUND(r.valor / 100.0, 2),
  'Entrada',
  r.tipo,
  r.status
FROM receipts r JOIN projects p ON p.id = r.project_id
WHERE r.deleted_at IS NULL AND r.linked_receipt_id IS NULL AND p.name = ?
"""


def load_from_db(path, project):
    """Lê um arquivo SQLite (.db baixado de prod via `flyctl ssh sftp get /data/dev.db`)
    e reproduz as transações com a lib sqlite3 embutida do Python — não precisa do binário
    sqlite3 (que NÃO existe na máquina de prod)."""
    import sqlite3
    con = sqlite3.connect(f"file:{path}?mode=ro", uri=True)
    try:
        rows = con.execute(DB_QUERY, (project, project)).fetchall()
    finally:
        con.close()
    if not rows:
        raise SystemExit(f"Nenhuma transação para o projeto {project!r}. "
                         f"Confira o nome (--project) ou o arquivo .db.")
    # idx na ordem das colunas do SELECT
    idx = {"origem": 2, "valor": 5, "tipo": 6, "status": 8, "categoria": 7, "desc": 4}
    return idx, rows


def load_transacoes(path):
    """Carrega a aba/linhas de Transações de um .xlsx OU de um .csv (sqlite3 -header -csv).
    Aceita as duas fontes pra cobrir export do app e dump direto do banco via Fly SSH."""
    if path.lower().endswith(".csv"):
        import csv
        with open(path, newline="", encoding="utf-8") as f:
            rows = [tuple(r) for r in csv.reader(f) if any((c or "").strip() for c in r)]
    else:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = "Transações" if "Transações" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet]
        rows = [r for r in ws.iter_rows(values_only=True) if r and any(c is not None for c in r)]
        wb.close()
    header = [(str(c).strip() if c is not None else "") for c in rows[0]]

    def col(*names):
        for n in names:
            if n in header:
                return header.index(n)
        raise SystemExit(f"Coluna não encontrada (esperava uma de {names}). Cabeçalho: {header}")

    idx = {
        "origem": col("Origem"),
        "valor": col("Valor", "Valor (R$)"),
        "tipo": col("Tipo"),
        "status": col("Status"),
        "categoria": col("Categoria"),
        "desc": col("Descrição", "Descricao", "Lançamento"),
    }
    data = rows[1:]
    return idx, data


def num(x):
    """Número robusto: aceita int/float (xlsx) ou string '-114.25' (csv)."""
    if isinstance(x, (int, float)):
        return x
    try:
        return float(str(x).strip())
    except (ValueError, TypeError):
        return 0.0


def analyze_prod(idx, data):
    g = lambda r, k: r[idx[k]]
    nonmanual = [r for r in data if g(r, "origem") != MANUAL]
    real = [r for r in nonmanual if g(r, "categoria") != NEUTRAL_CAT]
    neutras_rows = [r for r in nonmanual if g(r, "categoria") == NEUTRAL_CAT]

    recebimentos_reais = sum(num(g(r, "valor")) for r in real if num(g(r, "valor")) > 0)
    saidas_reais = sum(num(g(r, "valor")) for r in real if num(g(r, "valor")) < 0)
    neutras = sum(num(g(r, "valor")) for r in neutras_rows)

    realized = [r for r in real if g(r, "status") in REALIZED]
    ent_realiz = sum(num(g(r, "valor")) for r in realized if num(g(r, "valor")) > 0)
    sai_realiz = sum(num(g(r, "valor")) for r in realized if num(g(r, "valor")) < 0)
    liquido_realizado = ent_realiz + sai_realiz
    saldo_projetado = recebimentos_reais + saidas_reais

    # §10: caixa = saldo inicial + Σ lançamentos REALIZADOS da CONTA (inclui fatura paga pela conta).
    conta_realizado = sum(
        num(g(r, "valor")) for r in nonmanual
        if g(r, "origem") == CONTA and g(r, "status") in REALIZED
    )

    # Sanidade: resgates devem ser ENTRADA (valor > 0).
    resgates_negativos = [
        r for r in nonmanual
        if RESGATE_RE.search(str(g(r, "desc")) or "")
        and not EST_RESG_RE.search(str(g(r, "desc")) or "")
        and num(g(r, "valor")) < 0
    ]

    return {
        "metrics": {
            "recebimentos_reais": recebimentos_reais,
            "saidas_reais": saidas_reais,
            "neutras": neutras,
            "liquido_realizado": liquido_realizado,
            "saldo_projetado": saldo_projetado,
        },
        "conta_realizado": conta_realizado,
        "n_manuais": sum(1 for r in data if g(r, "origem") == MANUAL),
        "n_total": len(data),
        "origens": Counter(g(r, "origem") for r in data),
        "resgates_negativos": [(str(g(r, "desc"))[:35], num(g(r, "valor"))) for r in resgates_negativos],
    }


def targets_from_master(path):
    """Recalcula os alvos a partir do master (aba Resumo com valores em cache)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    flat = {}
    for r in wb["Resumo"].iter_rows(values_only=True):
        if r and r[0] and isinstance(r[1], (int, float)):
            flat[str(r[0]).strip()] = r[1]
    wb.close()
    return {
        "recebimentos_reais": flat.get("Entradas reais (pago)", DEFAULT_TARGETS["recebimentos_reais"]),
        "saidas_reais": flat.get("Saídas reais (pago)", DEFAULT_TARGETS["saidas_reais"]),
        "liquido_realizado": flat.get("Líquido realizado", DEFAULT_TARGETS["liquido_realizado"]),
        "saldo_projetado": flat.get("(=) Saldo futuro projetado", DEFAULT_TARGETS["saldo_projetado"]),
        "neutras": DEFAULT_TARGETS["neutras"],
        "caixa_conta": flat.get("(=) Saldo da conta hoje (bate com o app do banco)", DEFAULT_TARGETS["caixa_conta"]),
    }


def fmt(v):
    return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def main():
    ap = argparse.ArgumentParser(description="Reconciliação PESSOAL — prod vs master")
    src = ap.add_mutually_exclusive_group(required=True)
    src.add_argument("--db", help="arquivo SQLite baixado de prod (flyctl ssh sftp get /data/dev.db)")
    src.add_argument("--prod", help="export de produção (.xlsx ou .csv, com aba/colunas de Transações)")
    ap.add_argument("--project", default=DEFAULT_PROJECT, help="nome do projeto PESSOAL (modo --db)")
    ap.add_argument("--master", help="Consolidacao_Financeira.xlsx (recalcula os alvos)")
    ap.add_argument("--opening", type=float, default=DEFAULT_OPENING, help="saldo inicial da conta (R$)")
    ap.add_argument("--opening-date", default="2025-12-31")
    ap.add_argument("--bank", type=float, default=DEFAULT_BANK, help="saldo do banco p/ referência (R$)")
    args = ap.parse_args()

    targets = targets_from_master(args.master) if args.master else dict(DEFAULT_TARGETS)

    idx, data = load_from_db(args.db, args.project) if args.db else load_transacoes(args.prod)
    res = analyze_prod(idx, data)
    m = res["metrics"]
    caixa = args.opening + res["conta_realizado"]

    print("=" * 64)
    print("RECONCILIAÇÃO PESSOAL — produção vs master")
    print("=" * 64)
    print(f"Fonte: {args.db + f' (projeto: {args.project})' if args.db else args.prod}")
    print(f"Linhas: {res['n_total']}  |  Origens: {dict(res['origens'])}")
    print(f"Manuais excluídos: {res['n_manuais']}")
    print()

    rows = [
        ("Recebimentos reais", m["recebimentos_reais"], targets["recebimentos_reais"]),
        ("Saídas reais", m["saidas_reais"], targets["saidas_reais"]),
        ("Neutras (pgto fatura)", m["neutras"], targets["neutras"]),
        ("Líquido realizado", m["liquido_realizado"], targets["liquido_realizado"]),
        ("Saldo projetado", m["saldo_projetado"], targets["saldo_projetado"]),
        (f"Caixa §10 (ini {fmt(args.opening)})", caixa, targets["caixa_conta"]),
    ]
    ok_all = True
    print(f"{'Métrica':28} {'prod':>16} {'master':>16}  {'':4}")
    print("-" * 68)
    for label, got, tgt in rows:
        ok = abs(got - tgt) <= TOL
        ok_all = ok_all and ok
        print(f"{label:28} {fmt(got):>16} {fmt(tgt):>16}  {'OK ' if ok else 'FALHA'}")
    print("-" * 68)

    # Referência do banco (só informativo: §10 vs extrato).
    diff_bank = caixa - args.bank
    print(f"\nCaixa §10 vs banco ({fmt(args.bank)}): diferença {fmt(diff_bank)} "
          f"({'ok, arredondamento' if abs(diff_bank) <= 1 else 'CONFERIR'})")

    # Sanidade: resgates negativos = bug de import (deveriam ser ENTRADA).
    if res["resgates_negativos"]:
        print(f"\n⚠️  {len(res['resgates_negativos'])} RESGATE(s) com sinal NEGATIVO (deveriam ser entrada):")
        for desc, val in res["resgates_negativos"]:
            print(f"     {fmt(val):>14}  {desc}")
        ok_all = False
    else:
        print("\n✓ Resgates: todos como entrada (sem inversão de sinal).")

    print("\n" + ("✅ TUDO BATE." if ok_all else "❌ HÁ DIVERGÊNCIAS — ver acima."))
    sys.exit(0 if ok_all else 1)


if __name__ == "__main__":
    main()
